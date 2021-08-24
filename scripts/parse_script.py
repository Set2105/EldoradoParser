import json

from os import path

from time import sleep
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
from openpyxl.utils import get_column_letter

import win32api
import win32con
import cv2
import numpy as np
import pyscreenshot as ImageGrab
import pyautogui
import wx, ctypes

import re


def click(x, y):
    win32api.SetCursorPos((x, y))
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x, y, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x, y, 0, 0)


def find_patt(image, patt, thres):
    img_grey = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    patt_H, patt_W = patt.shape[:2]
    res = cv2.matchTemplate(img_grey, patt, cv2.TM_CCOEFF_NORMED)
    loc = np.where(res > thres)
    return patt_H, patt_W, list(zip(*loc[::-1]))


def console_log(message):
    now = datetime.now()
    print(f'{now.hour}:{now.minute}:{now.second}: {message}')


class Cell:
    row = 0
    cell = 0

    def __init__(self, row: int = 0, cell: int = 0):
        self.row = row
        self.cell = cell

    def __call__(self, a: int = 0, b: int = 0):
        return get_column_letter(self.cell+b) + str(self.row + a)

    def __repr__(self):
        return str(get_column_letter(self.cell) + str(self.row))

    def add(self, a: int, b: int):
        self.row += a
        self.cell += b


class EldoradoParser(webdriver.Chrome):
    authenticated = False
    current_url = ''
    last_url = ''

    success_buys = 0
    needed_success_buys = 3

    info_list = []
    info = {}
    info_id = 0
    goods_list = []

    def __init__(self):
        options = Options()
        options.add_argument('--disable-features=NetworkService')
        options.add_argument("--disable-notifications")
        options.add_argument('--enable-features=NetworkServiceWindowsSandbox')
        options.add_argument("--disable-infobars")
        chrome_path = path.abspath(path.join(path.dirname(__file__), 'chromedriver.exe'))
        with open('../options/contact_info.json', encoding='utf-8') as contact_json:
            self.info_list = json.load(contact_json)
            self.info = self.info_list[self.info_id]
        super().__init__(chrome_path, options=options)

    def get_url(self, url):
        self.last_url = self.current_url
        self.current_url = url
        return self.get(url)

    def login(self):
        self.get('https://www.eldorado.ru/auth.php')
        while True:
            char = input('Войдите в свой профиль и введите \'y\':\n')
            if char == 'y':
                break

    def login_mail(self, email, password):
        self.get('https://www.eldorado.ru/auth.php')
        self.find_element(By.CLASS_NAME, 'js-login-w-email').click()
        self.find_element(By.ID, 'USER_LOGIN').send_keys(email)
        self.find_element(By.ID, 'USER_PASSWORD').send_keys(password, Keys.ENTER)
        sleep(5)
        if self.current_url == 'https://www.eldorado.ru/personal/club/operations/':
            self.authenticated = True
        if self.last_url:
            self.get(self.current_url)

    def logout(self):
        self.get('https://www.eldorado.ru/personal/club/operations/?logout=yes')
        self.get(self.current_url)

    def find_buy_button(self):
        try:
            return self.find_element(By.CLASS_NAME, 'addToCartBig')
        except Exception as e:
            print(e)
            return None

    def add_item_to_basket(self):
        buy_button = self.find_buy_button()
        if buy_button is not None:
            self.mouse_click_element(By.CLASS_NAME, 'addToCartBig', '../buttons_scr/add_to_basket.bmp')
            return True
        else:
            return False

    @staticmethod
    def mouse_click_picture(button_picture):

        app = wx.PySimpleApp()

        SM_XVIRTUALSCREEN = 76
        SM_YVIRTUALSCREEN = 77
        SM_CXVIRTUALSCREEN = 78
        SM_CYVIRTUALSCREEN = 79

        user32 = ctypes.windll.user32
        width, height = user32.GetSystemMetrics(SM_CXVIRTUALSCREEN), user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)
        x, y = user32.GetSystemMetrics(SM_XVIRTUALSCREEN), user32.GetSystemMetrics(SM_YVIRTUALSCREEN)

        screen = wx.ScreenDC()
        bmp = wx.EmptyBitmap(width, height)
        mem = wx.MemoryDC(bmp)
        mem.Blit(0, 0, width, height, screen, x, y)
        del mem

        screenshot = ImageGrab.grab()
        img = np.array(screenshot.getdata(), dtype='uint8').reshape((screenshot.size[1], screenshot.size[0], 3))

        patt = cv2.imread(button_picture, 0)
        h, w, points = find_patt(img, patt, 0.60)
        pyautogui.FAILSAFE = False
        if len(points) != 0:
            pyautogui.moveTo(points[0][0] + w / 2, points[0][1] + h / 2)
            pyautogui.click()

    def move_to_next_button(self):
        action = ActionChains(self)
        element = self.find_element(By.CLASS_NAME, 'cartTotalPart')
        try:
            action.move_to_element(element).perform()
        except Exception as e:
            print(e)

    def input_person_info(self, By_obj, class_name, value):
        form = self.find_element(By_obj, class_name)
        for _ in range(25):
            form.send_keys(Keys.BACK_SPACE)
        if value:
            form.send_keys(value)
            return form

    @staticmethod
    def confirm_dropdown(element):
        if element is not None:
            sleep(3)
            element.send_keys(Keys.DOWN, Keys.ENTER)

    def fill_contact_info(self):
        self.input_person_info(By.ID, 'USER_FACE_NAME', self.info['Имя'])
        self.input_person_info(By.ID, 'USER_FACE_LAST_NAME', self.info['Фамилия'])
        self.input_person_info(By.ID, 'phone_mobile_1', self.info['Телефон'])
        self.input_person_info(By.ID, 'USER_EMAIL', self.info['Почта'])
        self.confirm_dropdown(self.input_person_info(By.ID, 'metro_dropdown', self.info['Метро']))
        self.confirm_dropdown(self.input_person_info(By.ID, 'ORDER_PROP_5_VALUE', self.info['Улица']))
        self.input_person_info(By.ID, 'address_house', self.info['Дом'])
        self.input_person_info(By.ID, 'address_building', self.info['Стоение'])
        self.input_person_info(By.ID, 'address_housing', self.info['Корпус'])
        self.input_person_info(By.ID, 'address_porch', self.info['Подьезд'])
        self.input_person_info(By.ID, 'address_floor', self.info['Этаж'])
        self.input_person_info(By.ID, 'address_room', self.info['Квартира'])
        self.input_person_info(By.ID, 'address_doorcode', self.info['Домофон'])

    def wait_located_element(self, By_obj, value, wait_time=10):
        return WebDriverWait(self, wait_time).until(EC.presence_of_element_located((By_obj, value)))

    def mouse_click_element(self, By_obj, value, element_picture):
        if self.wait_located_element(By_obj, value):
            self.mouse_click_picture(element_picture)

    def make_basket_order(self):
        try:
            self.get('https://www.eldorado.ru/personal/basket.php')
            self.move_to_next_button()
            self.mouse_click_element(By.ID, 'delivery_radio_rc13906-styler', '../buttons_scr/pointer.bmp')
            self.move_to_next_button()
            self.mouse_click_element(By.CLASS_NAME, 'successBttnCP', '../buttons_scr/next_button.bmp')
            self.fill_contact_info()
            delivery_date = self.find_element(By.CLASS_NAME, 'q-delivery-date-value').text
            self.mouse_click_element(By.CLASS_NAME, 'addToCartBig', '../buttons_scr/next_button2.bmp')
            self.mouse_click_element(By.ID, 'delivery-box', '../buttons_scr/payment.bmp')
            self.mouse_click_picture('../buttons_scr/next_button2.bmp')
            self.mouse_click_element(By.CLASS_NAME, 'orderConfirmSubmitButton', '../buttons_scr/Accept.bmp')
            WebDriverWait(self, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'text_green')))
            order_num = self.find_element(By.CLASS_NAME, 'text_green').text
            console_log(f'Order {order_num}({delivery_date}) was made')
            return {'order_num': order_num, 'delivery_date': delivery_date}
        except Exception as e:
            print(e)
            return None

    def clear_basket(self):
        self.get('https://www.eldorado.ru/personal/basket.php')
        elements = self.find_elements(By.CLASS_NAME, 'basketBlockRow')
        for element in elements:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "q-basketBlockClouser-button"))).click()
            except:
                pass

    def check_and_buy_good(self, url, good_name):
        self.get_url(url)
        if self.add_item_to_basket():
            return self.make_basket_order()
        else:
            console_log(f'{good_name} is not sellable')

    def save_xlsx(self, data, sheet_name, link):
        try:
            wb = openpyxl.load_workbook('../result.xlsx')
        except Exception as e:
            print(e)
            wb = openpyxl.Workbook()

        try:
            ws = wb[sheet_name]
        except Exception as e:
            print(e)
            ws = wb.create_sheet(sheet_name)

        writing_cell = Cell(1, 1)
        if ws[writing_cell()].value is None:
            ws[writing_cell()].value = sheet_name
            ws[writing_cell(0, 1)].value = link
        while ws[writing_cell()].value is not None:
            writing_cell.add(1, 0)

        order = data['order_num']
        delivery_date = data['delivery_date']
        address = '{} д{}'.format(self.info['Улица'], self.info['Дом'])

        ws[writing_cell()].value = order
        ws[writing_cell(0, 1)].value = delivery_date
        ws[writing_cell(0, 2)].value = address

        wb.save('../result.xlsx')
        console_log(f'Order {order} was saved to result.xlsx')

    def save(self, data, good_id):
        self.goods_list[good_id]['Количество'] -= 1
        self.save_xlsx(data, self.goods_list[good_id]['Имя'], self.goods_list[good_id]['Ссылка'])
        if self.goods_list[good_id]['Количество'] < 1:
            del self.goods_list[good_id]

    def change_address_after_success_buy(self):
        self.success_buys += 1
        if self.success_buys == self.needed_success_buys:
            self.success_buys = 0
            self.info_id += 1

            if self.info_id >= len(self.info_list):
                self.info_id = 0

            self.info = self.info_list[self.info_id]
            console_log('Новый адрес: {self.info}')

    def loop_condition(self):
        for good in self.goods_list:
            if good['Количество'] > 0:
                return True
        return False

    def load_goods_list(self):
        with open('../goods_list.txt', encoding="utf-8") as file:
            console_log('loading goods list')
            for string in file:
                data = re.split(' ', string)
                try:
                    self.goods_list.append({
                        'Имя': data[0],
                        'Ссылка': data[1],
                        'Количество': int(data[2])})
                except Exception as e:
                    print(e)
            console_log('goods_list loaded')

    def start_loop(self, timer):
        while self.loop_condition():
            for i in range(0, len(self.goods_list)):
                good = self.goods_list[i]
                while good['Количество'] > 0:
                    self.clear_basket()
                    data = self.check_and_buy_good(good['Ссылка'], good['Имя'])
                    if data is None:
                        break
                    else:
                        self.change_address_after_success_buy()
                        self.save(data, i)
            console_log(f'New try in {timer} sec')
            sleep(timer-10)
            console_log('Start in 10 sec')
            for i in range(0, 10):
                sleep(1)
                console_log(f'{10-i}')


driver = EldoradoParser()
driver.login()
driver.load_goods_list()
driver.start_loop(20)
driver.close()



